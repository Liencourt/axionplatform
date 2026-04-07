import logging
import threading
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from .extractor import processar_encarte
from .models import Concorrente, ExtraçãoEncarte, ProdutoEncarte

logger = logging.getLogger('encartes')


# ─── Lista de Extrações ───────────────────────────────────────────────────────

@login_required
def lista_extracoes(request):
    extracoes = (
        ExtraçãoEncarte.objects
        .filter(empresa=request.empresa)
        .select_related('concorrente')
        .order_by('-data_extracao')
    )

    # Filtros opcionais
    concorrente_id = request.GET.get('concorrente')
    if concorrente_id:
        extracoes = extracoes.filter(concorrente_id=concorrente_id)

    paginator = Paginator(extracoes, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    concorrentes = Concorrente.objects.filter(empresa=request.empresa, ativo=True)

    return render(request, 'encartes/lista.html', {
        'page_obj': page_obj,
        'concorrentes': concorrentes,
        'concorrente_selecionado': concorrente_id,
    })


# ─── Upload e Disparo de Extração ────────────────────────────────────────────

@login_required
def upload_encarte(request):
    concorrentes = Concorrente.objects.filter(empresa=request.empresa, ativo=True)

    if request.method == 'POST':
        concorrente_id = request.POST.get('concorrente_id')
        novo_concorrente = request.POST.get('novo_concorrente', '').strip()
        programa_fidelidade = request.POST.get('programa_fidelidade', '').strip() or None
        vigencia_inicio = request.POST.get('vigencia_inicio')
        vigencia_fim = request.POST.get('vigencia_fim')
        arquivo = request.FILES.get('arquivo_pdf')

        if not arquivo:
            messages.error(request, 'Selecione um arquivo PDF para continuar.')
            return render(request, 'encartes/upload.html', {'concorrentes': concorrentes})

        if not arquivo.name.lower().endswith('.pdf'):
            messages.error(request, 'O arquivo deve ser um PDF.')
            return render(request, 'encartes/upload.html', {'concorrentes': concorrentes})

        if not vigencia_inicio or not vigencia_fim:
            messages.error(request, 'Informe as datas de vigência das ofertas.')
            return render(request, 'encartes/upload.html', {'concorrentes': concorrentes})

        # Resolver concorrente
        if novo_concorrente:
            concorrente, _ = Concorrente.objects.get_or_create(
                empresa=request.empresa,
                nome=novo_concorrente,
                defaults={'programa_fidelidade': programa_fidelidade},
            )
        elif concorrente_id:
            concorrente = get_object_or_404(Concorrente, pk=concorrente_id, empresa=request.empresa)
        else:
            messages.error(request, 'Selecione um concorrente ou informe o nome de um novo.')
            return render(request, 'encartes/upload.html', {'concorrentes': concorrentes})

        try:
            extracao = ExtraçãoEncarte.objects.create(
                empresa=request.empresa,
                concorrente=concorrente,
                arquivo_pdf=arquivo,
                vigencia_inicio=vigencia_inicio,
                vigencia_fim=vigencia_fim,
            )
        except Exception as exc:
            logger.exception("Erro ao criar extração: %s", exc)
            messages.error(request, f'Erro ao salvar o arquivo: {exc}')
            return render(request, 'encartes/upload.html', {'concorrentes': concorrentes})

        # Disparar extração em thread separada para não bloquear o request
        thread = threading.Thread(
            target=_executar_extracao_background,
            args=(extracao.pk,),
            daemon=True,
        )
        thread.start()

        messages.success(
            request,
            f'Encarte enviado! A extração de "{concorrente.nome}" foi iniciada. '
            'Aguarde alguns minutos e atualize a página para ver o resultado.'
        )
        return redirect('detalhe_extracao', extracao_id=extracao.pk)

    return render(request, 'encartes/upload.html', {'concorrentes': concorrentes})


def _executar_extracao_background(extracao_id: int) -> None:
    """Wrapper para rodar o extractor em background thread."""
    try:
        processar_encarte(extracao_id)
    except Exception as exc:
        logger.exception("Falha na extração background id=%d: %s", extracao_id, exc)


# ─── Detalhe de uma Extração ──────────────────────────────────────────────────

@login_required
def detalhe_extracao(request, extracao_id):
    extracao = get_object_or_404(
        ExtraçãoEncarte,
        pk=extracao_id,
        empresa=request.empresa,
    )

    produtos_qs = (
        ProdutoEncarte.objects
        .filter(extracao=extracao)
        .prefetch_related('precos')
        .order_by('pagina', 'nome')
    )

    # Filtros
    categoria = request.GET.get('categoria', '').strip()
    busca = request.GET.get('q', '').strip()

    if categoria:
        produtos_qs = produtos_qs.filter(categoria=categoria)
    if busca:
        produtos_qs = produtos_qs.filter(nome__icontains=busca)

    categorias = (
        ProdutoEncarte.objects
        .filter(extracao=extracao)
        .exclude(categoria__isnull=True)
        .exclude(categoria='')
        .values_list('categoria', flat=True)
        .distinct()
        .order_by('categoria')
    )

    paginator = Paginator(produtos_qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'encartes/detalhe.html', {
        'extracao': extracao,
        'page_obj': page_obj,
        'categorias': categorias,
        'categoria_selecionada': categoria,
        'busca': busca,
    })


# ─── Status (polling AJAX) ────────────────────────────────────────────────────

@login_required
def status_extracao(request, extracao_id):
    extracao = get_object_or_404(ExtraçãoEncarte, pk=extracao_id, empresa=request.empresa)
    return JsonResponse({
        'status': extracao.status,
        'status_display': extracao.get_status_display(),
        'total_produtos': extracao.total_produtos,
        'total_precos': extracao.total_precos,
        'erro_mensagem': extracao.erro_mensagem or '',
    })


# ─── Excluir Extração ─────────────────────────────────────────────────────────

@login_required
@require_POST
def excluir_extracao(request, extracao_id):
    extracao = get_object_or_404(ExtraçãoEncarte, pk=extracao_id, empresa=request.empresa)
    nome = str(extracao)
    extracao.delete()
    messages.success(request, f'Extração "{nome}" excluída com sucesso.')
    return redirect('lista_extracoes')


# ─── Exportar Excel ──────────────────────────────────────────────────────────

@login_required
def exportar_extracao_excel(request, extracao_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    extracao = get_object_or_404(ExtraçãoEncarte, pk=extracao_id, empresa=request.empresa)

    produtos_qs = (
        ProdutoEncarte.objects
        .filter(extracao=extracao)
        .prefetch_related('precos')
        .order_by('pagina', 'nome')
    )

    categoria = request.GET.get('categoria', '').strip()
    busca = request.GET.get('q', '').strip()
    if categoria:
        produtos_qs = produtos_qs.filter(categoria=categoria)
    if busca:
        produtos_qs = produtos_qs.filter(nome__icontains=busca)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Produtos'

    # Cabeçalho
    headers = ['Produto', 'Marca', 'Categoria', 'Quantidade', 'Condição Especial',
               'Preço Normal', 'Melhor Preço', 'Tipo Melhor Preço', 'Condição Preço', 'Página']
    header_fill = PatternFill(start_color='1d4ed8', end_color='1d4ed8', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Larguras
    col_widths = [45, 20, 18, 14, 30, 14, 14, 18, 30, 8]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Dados
    TIPO_LABEL = {'clube': 'Clube', 'promocional': 'Promo', 'condicional': 'Condicional', 'normal': 'Normal'}

    for row_idx, produto in enumerate(produtos_qs, start=2):
        precos = list(produto.precos.all())
        preco_normal = next((p for p in precos if p.tipo == 'normal'), None)
        melhor = next(
            (p for p in precos if p.tipo == 'clube'), None
        ) or next(
            (p for p in precos if p.tipo == 'promocional'), None
        ) or next(
            (p for p in precos if p.tipo == 'condicional'), None
        ) or preco_normal

        ws.append([
            produto.nome,
            produto.marca or '',
            (produto.categoria or '').capitalize(),
            produto.quantidade or '',
            produto.condicao_especial or '',
            float(preco_normal.valor) if preco_normal else '',
            float(melhor.valor) if melhor and melhor != preco_normal else '',
            TIPO_LABEL.get(melhor.tipo, melhor.tipo) if melhor and melhor != preco_normal else '',
            melhor.condicao or '' if melhor and melhor != preco_normal else '',
            produto.pagina,
        ])

        # Destaque para linhas com preço clube
        if melhor and melhor.tipo == 'clube':
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color='EFF6FF', end_color='EFF6FF', fill_type='solid'
                )

    # Aba de metadados
    ws_meta = wb.create_sheet('Informações')
    ws_meta.append(['Campo', 'Valor'])
    ws_meta.append(['Concorrente', extracao.concorrente.nome])
    ws_meta.append(['Vigência', extracao.vigencia_str])
    ws_meta.append(['Data da extração', extracao.data_extracao.strftime('%d/%m/%Y %H:%M')])
    ws_meta.append(['Total de produtos', extracao.total_produtos])
    ws_meta.append(['Total de preços', extracao.total_precos])
    ws_meta.append(['Preços clube/fidelidade', extracao.precos_clube])
    ws_meta.append(['Preços promocionais', extracao.precos_promocional])
    ws_meta.append(['Modelo de extração', extracao.modelo_extracao])
    ws_meta.column_dimensions['A'].width = 28
    ws_meta.column_dimensions['B'].width = 35

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"encarte_{extracao.concorrente.nome.lower().replace(' ', '_')}_{extracao.vigencia_inicio}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Gerenciar Concorrentes ───────────────────────────────────────────────────

@login_required
def lista_concorrentes(request):
    concorrentes = Concorrente.objects.filter(empresa=request.empresa).order_by('nome')
    return render(request, 'encartes/concorrentes.html', {'concorrentes': concorrentes})


@login_required
@require_POST
def excluir_concorrente(request, concorrente_id):
    concorrente = get_object_or_404(Concorrente, pk=concorrente_id, empresa=request.empresa)
    nome = concorrente.nome
    concorrente.delete()
    messages.success(request, f'Concorrente "{nome}" excluído.')
    return redirect('lista_concorrentes')
